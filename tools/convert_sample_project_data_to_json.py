from __future__ import annotations

import csv
import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
PROJECT_DIR = ROOT / "sample_project"


def clean_number(value: Any) -> int | float:
    number = float(value)
    if number.is_integer():
        return int(number)
    return number


def read_load_points() -> list[dict[str, Any]]:
    load_points = []
    with (PROJECT_DIR / "Belastinglocaties.csv").open(newline="", encoding="utf-8") as file:
        reader = csv.reader(file, delimiter=",")
        for row in reader:
            if not row:
                continue
            load_point_id, x_mm, y_mm, design_load_kn = row[:4]
            load_points.append(
                {
                    "id": int(load_point_id),
                    "name": f"Load point {load_point_id}",
                    "x_mm": clean_number(x_mm),
                    "y_mm": clean_number(y_mm),
                    "design_load_kn": clean_number(design_load_kn),
                }
            )
    return load_points


def read_cpts() -> list[dict[str, Any]]:
    workbook = load_workbook(PROJECT_DIR / "Sonderingen.xlsx", read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    cpts = []
    for row in worksheet.iter_rows(values_only=True):
        if not row or row[0] is None:
            continue
        cpt_id, x_mm, y_mm = row[:3]
        cpts.append(
            {
                "id": int(cpt_id),
                "name": f"CPT {int(cpt_id)}",
                "x_mm": clean_number(x_mm),
                "y_mm": clean_number(y_mm),
            }
        )
    return cpts


def read_bearing_capacities() -> list[dict[str, Any]]:
    workbook = load_workbook(
        PROJECT_DIR / "Draagvermogens.xlsx",
        read_only=True,
        data_only=True,
    )
    worksheet = workbook[workbook.sheetnames[0]]
    capacities = []
    rows = worksheet.iter_rows(values_only=True)
    next(rows, None)
    for row in rows:
        if not row or row[0] is None:
            continue
        cpt_id, pile_tip_level_m, pile_size_mm, frd_kn = row[:4]
        if frd_kn is None:
            continue
        capacities.append(
            {
                "cpt_id": int(cpt_id),
                "pile_tip_level_m": clean_number(pile_tip_level_m),
                "pile_size_mm": clean_number(pile_size_mm),
                "frd_kn": clean_number(frd_kn),
            }
        )
    return capacities


def write_json(filename: str, items: list[dict[str, Any]]) -> None:
    payload = {
        "schema_version": 1,
        "items": items,
    }
    path = PROJECT_DIR / filename
    path.write_text(
        json.dumps(payload, indent=2, ensure_ascii=False) + "\n",
        encoding="utf-8",
    )


def main() -> None:
    write_json("load_points.json", read_load_points())
    write_json("cpts.json", read_cpts())
    write_json("bearing_capacities.json", read_bearing_capacities())


if __name__ == "__main__":
    main()
